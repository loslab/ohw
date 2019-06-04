# -*- coding: utf-8 -*-

import matplotlib.pyplot as plt
import numpy as np
import pathlib
from scipy.signal import argrelextrema
from PyQt5.QtWidgets import QSizePolicy
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from openpyxl import Workbook
from openpyxl.styles import Font

class PeakDetection():
    def __init__(self):
        self.motion = None
        self.timeindex = None
        self.Peaks = []
        self.hipeaks = []
        self.lopeaks = []
        
        """
        self.peaktimes_th = None  #times of thresholded peaks
        self.peak_heights_th = None  # peakheights of thresholded peaks
        self.sorted_peaks = None  # list of sorted peaks (high/low) and corresponding times
        self.peakstatistics = None # list of peakstatistics: max_contraction, std, max_relaxation, std
        self.time_intervals = None # list of calculated intervals and std: contraction, relaxation, contr-rel
        self.bpm = None # list of bpm + std
        self.foundPeaks = 0  # number of found peaks
        """
    
    def set_data(self, timeindex, mean_absMotions):
        """
            reads in x,y datapairs (timeindex, mean_absMotions) and resets all output dicts
        """
        self.motion = mean_absMotions # rename to motion only, can be every motion
        #self.mean_absMotions = mean_absMotions
        self.timeindex = timeindex
        self.Peaks = []
        self.hipeaks = [] #needed here ageain? or just call orderPeaks?
        self.lopeaks = []
        
        """
        self.foundPeaks = 0
        
        self.peaktimes_th = None
        self.peak_heights_th = None
        
        self.sorted_peaks = {"t_peaks_low_sorted":None,"peaks_low_sorted":None,"t_peaks_high_sorted":None,"peaks_high_sorted":None}
        self.peakstatistics = {"max_contraction":None, "max_contraction_std": None, "max_relaxation": None, "max_relaxation_std": None}

        self.time_intervals = {"contraction_interval_mean":None, "contraction_interval_std":None, "relaxation_interval_mean":None, "relaxation_interval_std":None, "contr_relax_interval_mean":None, "contr_relax_interval_std":None}
        self.bpm = {"bpm_mean":None, "bpm_std": None}
        """
    
    def set_peaks(self,Peaks):
        self.Peaks = Peaks
        
    def detect_peaks(self, ratio, number_of_neighbours):
        
        print("detecting peaks with ratio: ", ratio, " and neighbours: ", number_of_neighbours)
        
        peaks_ = argrelextrema(self.motion, np.greater, order=number_of_neighbours)
        peaks_ = np.array(peaks_[0])  #location of peaks in inputarray
        
        #remove peaks that are smaller than a certain threshold, e.g. ratio = 1/15
        peakheights = self.motion[peaks_]
        threshold = ratio * np.max(peakheights)
        peaks_th = peaks_[peakheights > threshold]  #th=thresholded?
        
        foundPeaks = peaks_th.shape[0] #set number of found peaks
        print(foundPeaks, " peaks found")

        self.Peaks = sorted(list(peaks_th))
    
    def order_peaks(self):
        '''order high/low peaks, return index'''
        if len(self.Peaks) == 0:
            self.hipeaks, self.lopeaks = [],[]
            return
        
        # just order alternating, start with highest peak
        peakheights = self.motion[self.Peaks]
        highest_peak = np.argmax(peakheights)
        
        even = self.Peaks[::2]
        odd = self.Peaks[1::2]
        if highest_peak %2 == 0:    # if highest peak in even index
            hipeaks = even
            lopeaks = odd
        else:
            hipeaks = odd
            lopeaks = even
        self.hipeaks, self.lopeaks = hipeaks, lopeaks
    
    def calc_peakstatistics(self):
        decimals = 2
        motion_high = self.motion[self.hipeaks] if self.hipeaks != None else []
        motion_low = self.motion[self.lopeaks] if self.lopeaks != None else []
        times_high = self.timeindex[self.hipeaks] if self.hipeaks != None else []
        times_low = self.timeindex[self.lopeaks] if self.lopeaks != None else []

        max_contraction = np.round(np.mean(motion_high), decimals)
        max_contraction_std = np.round(np.std(motion_high),  decimals)        
        max_relaxation      = np.round(np.mean(motion_low),  decimals)
        max_relaxation_std  = np.round(np.std(motion_low),   decimals)
        
        contraction_intervals = np.diff(np.sort(times_high))
        relaxation_intervals = np.diff(np.sort(times_low))
        
        contraction_interval_mean = np.round(np.mean(contraction_intervals), decimals)
        contraction_interval_std = np.round(np.std(contraction_intervals), decimals)
        relaxation_interval_mean = np.round(np.mean(relaxation_intervals), decimals)
        relaxation_interval_std = np.round(np.std(relaxation_intervals), decimals)    
        
        contr_relax_intervals = []
        
        if len(self.Peaks) >= 2: # calculate contr_relax only if 1 relax peak after 1 contr peak detected
            shift = 0 if times_high[0] < times_low[0] else 1
            for low_peak, high_peak in zip(times_low[shift:], times_high):
                # depending on which peak is first the true interval might be from the next peak on....
                time_difference = low_peak - high_peak           
                contr_relax_intervals.append(time_difference)
        contr_relax_interval_mean = np.round(np.mean(np.asarray(contr_relax_intervals)), decimals)
        contr_relax_interval_std = np.round(np.std(np.asarray(contr_relax_intervals)), decimals)            

        bpm_mean = np.round(np.mean(60/contraction_intervals), decimals)
        bpm_std = np.round(np.std(60/contraction_intervals), decimals)        
        
        peakstatistics = {"max_contraction":max_contraction, "max_contraction_std":max_contraction_std,
                "max_relaxation":max_relaxation, "max_relaxation_std":max_relaxation_std,
                "contraction_interval_mean":contraction_interval_mean, "contraction_interval_std":contraction_interval_std,
                "relaxation_interval_mean":relaxation_interval_mean, "relaxation_interval_std":relaxation_interval_std,
                "bpm_mean":bpm_mean, "bpm_std": bpm_std,
                "contr_relax_interval_mean":contr_relax_interval_mean, "contr_relax_interval_std":contr_relax_interval_std}
        
        self.peakstatistics = peakstatistics
    
    def get_peakstatistics(self):
        return self.peakstatistics
    
    def export_analysis(self, results_folder):
        #save peaks and peakanalysis to excel file:
        
        workbook_peaks = Workbook()
        sheet_peaks = workbook_peaks.active
        sheet_peaks.title = 'Analyzed peaks'
        sheet_peaks.append(['t [s]', 'motion [µm/s]', 'high/low'])         
        
        for peak in self.Peaks:
        
            time = self.timeindex[peak]
            motion = self.motion[peak]
            
            if peak in self.hipeaks:
                sheet_peaks.append([time, motion, "high"])
            elif peak in self.lopeaks:
                sheet_peaks.append([time, motion, "low"])
        
        sheet_analysis = workbook_peaks.create_sheet("Peak analysis")
        #ws.title = "New Title"
        ps = self.peakstatistics # shorten variablename...
        
        sheet_analysis.append(['value', 'mean', 'standard deviation', 'unit'])
        sheet_analysis.append(['maximum contraction', ps["max_contraction"], ps["max_contraction_std"], u'\xb5m/s'])
        sheet_analysis.append(['maximum relaxation', ps["max_relaxation"], ps["max_relaxation_std"], u'\xb5m/s'])
        sheet_analysis.append(['mean contraction interval', ps["contraction_interval_mean"], ps["contraction_interval_std"], 's'])
        sheet_analysis.append(['mean relaxation interval', ps["relaxation_interval_mean"], ps["relaxation_interval_std"], 's'])
        sheet_analysis.append(['mean contraction-relaxation interval', ps["contr_relax_interval_mean"], ps["contr_relax_interval_std"], 's'])
        sheet_analysis.append(['beating rate', ps["bpm_mean"], ps["bpm_std"], 'beats/min'])        

        sheet_analysis.column_dimensions['A'].width = 31
        sheet_analysis.column_dimensions['C'].width = 17
        
        sheet_kinetics = workbook_peaks.create_sheet("Kinetics",0)
        sheet_kinetics.append(['t [s]', 'motion [µm/s]'])

        for time, motion in zip(self.timeindex,self.motion):
            sheet_kinetics.append([time, motion])
                
        save_file = str(results_folder / 'Motionanalysis.xlsx')
        workbook_peaks.save(save_file)

    """
    def exportStatistics(self, analysis_meta):
        font_bold = Font(bold=True)                     
        workbook_statistics = Workbook()
        
        sheet_peaks = workbook_statistics.active
        sheet_peaks.Name = 'Statistics'
        
        inputpath = analysis_meta["inputpath"]
        results_folder = analysis_meta["results_folder"]
        scalingfactor = analysis_meta["scalingfactor"]
        blockwidth = analysis_meta["MV_parameters"]["blockwidth"]
        delay = analysis_meta["MV_parameters"]["delay"]
        max_shift = analysis_meta["MV_parameters"]["max_shift"]
        
        #add the used parameters
        sheet_peaks.append(['Evaluated file/ folder: ', str(inputpath)])
        sheet_peaks.append(['Used parameters:', '', 'unit'])
        sheet_peaks.append(['scaling factor for calculation: ', scalingfactor, 'rel. to input'])
        sheet_peaks.append(['width of macroblocks ', blockwidth, 'pixels'])
        sheet_peaks.append(['delay between images', delay, 'frames'])
        #sheet_peaks.append(['framerate', fps, 'frames/sec']) #really needed?
        sheet_peaks.append(['maximum allowed movement ', max_shift, 'pixels'])
        
        #add the statistical results         
        sheet_peaks.append([''])
             
        #turn some cells'style to bold
        bold_cells = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11', 'A12', 'A13', 'A14', 'A15', 'A16', 'B10', 'C2','C10', 'D10']

        for cell in bold_cells:
            sheet_peaks[cell].font = font_bold
            
        #save the file 
        save_file = str(results_folder / 'Statistics.xlsx')
        workbook_statistics.save(save_file)
    """
    """
    def exportEKG_CSV(self, results_folder):
        font_bold = Font(bold=True) 
        
        workbook_ekg = Workbook()
        sheet_peaks_ekg = workbook_ekg.active
        sheet_peaks_ekg.Name = 'EKG values'
        sheet_peaks_ekg.append(['t [s]', 'v_mean [µm/s]'])

        for time, peak in zip(self.timeindex, self.mean_absMotions):
            sheet_peaks_ekg.append([time, peak])        
        
        #turn some cells'style to bold
        bold_cells = ['A1', 'B1']

        for cell in bold_cells:
            sheet_peaks_ekg[cell].font = font_bold
        
        #save peaks
        save_file = str(results_folder / 'EKG.xlsx')
        workbook_ekg.save(save_file)
    """