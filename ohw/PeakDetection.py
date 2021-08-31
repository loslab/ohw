# -*- coding: utf-8 -*-

import matplotlib.pyplot as plt
import numpy as np
import pathlib
#from scipy.signal import argrelextrema
from scipy.signal import find_peaks
from PyQt5.QtWidgets import QSizePolicy
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from openpyxl import Workbook
from openpyxl.styles import Font
from ohw import helpfunctions

class PeakDetection():
    def __init__(self):
        self.motion = None
        self.timeindex = None
        self.Peaks = [] # list of peaks = indices of timeindex/ motion array where peak occurs
        self.hipeaks = []
        self.lopeaks = []
        self.peakmode = "alternating"
        self.motiondescription = "Fluorescence Intensity [a.u.]" # used to label y-axis
        self.decimals = 2 # accuracy of results
    
    def set_data(self, timeindex, motion):
        """
            reads in x,y datapairs (timeindex, motion) and resets all output dicts
        """
        self.motion = motion # renamed to motion only, can be every 1D representation
        self.timeindex = timeindex
        self.Peaks = []
        self.hipeaks = [] #needed here ageain? or just call orderPeaks?
        self.lopeaks = []
    
    def set_peaks(self,Peaks):
        ''' assings list of Peakindices, sorts '''
        self.Peaks = sorted(Peaks)

    def get_savedata(self):
        savedict = {"motion": self.motion, "timeindex":self.timeindex, "Peaks":self.Peaks, "peakmode":self.peakmode, "motiondescription":self.motiondescription} # or init from postproc?
        return savedict

    def load_data(self, savedict):
        self.motion = savedict["motion"]
        self.timeindex = savedict["timeindex"]
        self.Peaks = savedict["Peaks"]
        self.peakmode = savedict["peakmode"]
        self.motiondescription = savedict["motiondescription"]
        
        self.assign_peaks()
        
    def detect_peaks(self, ratio, number_of_neighbours, prominence = 2):
        
        print("detecting peaks with ratio: ", ratio, " and neighbours: ", number_of_neighbours)
        
        #peaks_ = argrelextrema(self.motion, np.greater_equal, order=number_of_neighbours)
        peaks_ = find_peaks(self.motion, distance= number_of_neighbours, prominence = prominence)
        peaks_ = np.array(peaks_[0])  #location of peaks in inputarray
        
        #remove peaks that are smaller than a certain threshold, e.g. ratio = 1/15
        peakheights = self.motion[peaks_]
        threshold = ratio * np.max(peakheights)
        peaks_th = peaks_[peakheights > threshold]  #thresholded peaks
        
        foundPeaks = peaks_th.shape[0] #set number of found peaks
        print(foundPeaks, " peaks found")

        self.Peaks = sorted(list(peaks_th))
    
    def remove_peaks(self, tmin, tmax, vmax):
        """ removes peaks in range tmin->tmax which are smaller than vmax """
        
        # get adjacent frame corresponding to provided time
        Nmin = (np.abs(self.timeindex - tmin)).argmin()
        Nmax = (np.abs(self.timeindex - tmax)).argmin()
        
        # get indices of peaks in selected window
        peaksel = np.array(self.Peaks)
        peaksel = peaksel[np.where((peaksel>=Nmin) & (peaksel<=Nmax))]
        
        peakheights = self.motion[peaksel]
        delpeaks = peaksel[peakheights < vmax]
        
        self.Peaks = np.array(self.Peaks)
        self.Peaks = np.setdiff1d(self.Peaks, delpeaks)
        self.Peaks = list(self.Peaks)

    
    def set_peakmode(self, mode="alternating"):
        """ 
            sets mode how peaks are assigned to contraction/ relaxation
            - alternating
            - high
        """
        self.peakmode = mode

    def set_description(self, description):
        self.motiondescription = description
    
    def assign_peaks(self):
        ''' order high/low peaks, include index of each peak in each specific list '''
        if len(self.Peaks) == 0:
            self.hipeaks, self.lopeaks = [],[]
            return
        
        if self.peakmode == "alternating":
        
            # order alternating, start with highest peak
            peakheights = self.motion[self.Peaks]
            highest_peak = np.argmax(peakheights)
            
            even = self.Peaks[::2]
            odd = self.Peaks[1::2]
            if highest_peak %2 == 0:    # if highest peak in even index
                hipeaks = even
                lopeaks = odd
            else:
                hipeaks = odd
                lopeaks = even
        
        # all found peaks are high peaks
        elif self.peakmode == "high":
            lopeaks = []
            hipeaks = self.Peaks
            
        self.hipeaks, self.lopeaks = hipeaks, lopeaks
    
    def calc_peakstatistics(self):
        """ calculates statistics of found peaks like bpm,... """
   
        motion_high = self.motion[self.hipeaks] if self.hipeaks != None else [] #TODO: check, might be written easier
        motion_low = self.motion[self.lopeaks] if self.lopeaks != None else []
        times_high = self.timeindex[self.hipeaks] if self.hipeaks != None else []
        times_low = self.timeindex[self.lopeaks] if self.lopeaks != None else []

        max_contraction     = np.round(np.mean(motion_high), self.decimals)
        max_contraction_std = np.round(np.std(motion_high),  self.decimals)        
        max_relaxation      = np.round(np.mean(motion_low),  self.decimals)
        max_relaxation_std  = np.round(np.std(motion_low),   self.decimals)
        
        contraction_intervals = np.diff(np.sort(times_high))
        relaxation_intervals = np.diff(np.sort(times_low))
        
        contraction_interval_mean = np.round(np.mean(contraction_intervals), self.decimals)
        contraction_interval_std = np.round(np.std(contraction_intervals),   self.decimals)
        relaxation_interval_mean = np.round(np.mean(relaxation_intervals),   self.decimals)
        relaxation_interval_std = np.round(np.std(relaxation_intervals),     self.decimals)    

        bpm_mean = np.round(np.mean(60/contraction_intervals),         self.decimals)
        bpm_std = np.round(np.std(60/contraction_intervals), self.decimals)        
        
        self.peakstatistics = {"max_contraction":max_contraction, "max_contraction_std":max_contraction_std,
                "max_relaxation":max_relaxation, "max_relaxation_std":max_relaxation_std,
                "contraction_interval_mean":contraction_interval_mean, "contraction_interval_std":contraction_interval_std,
                "relaxation_interval_mean":relaxation_interval_mean, "relaxation_interval_std":relaxation_interval_std,
                "bpm_mean":bpm_mean, "bpm_std": bpm_std, "contr_relax_interval_mean":np.nan, "contr_relax_interval_std":np.nan}

        # calculate timings between alternating peaks
        if self.peakmode == "alternating":
            contr_relax_stats = self.get_contr_relax_stats(times_high, times_low)
            self.peakstatistics.update(contr_relax_stats)
    
    def get_contr_relax_stats(self, times_high, times_low):
        """
            calculates statistics based on times between contraction and relaxation
        """

        intervals = []
        
        if len(self.Peaks) >= 2: # calculate contr_relax only if 1 relax peak after 1 contr peak detected
            shift = 0 if times_high[0] < times_low[0] else 1
            for low_peak, high_peak in zip(times_low[shift:], times_high):
                # depending on which peak is first the true interval might be from the next peak on....
                time_difference = low_peak - high_peak           
                intervals.append(time_difference)
        contr_relax_interval_mean = np.round(np.mean(np.asarray(intervals)), self.decimals)
        contr_relax_interval_std = np.round(np.std(np.asarray(intervals)), self.decimals)
        
        return {"contr_relax_interval_mean":contr_relax_interval_mean, "contr_relax_interval_std":contr_relax_interval_std}
    
    def get_peakstatistics(self):
        return self.peakstatistics

    def calc_bpm_evolution(self, contraction_intervals):
        """
            calculates frequency evolution based from time between consecutive contraction peaks
        """
        
        freqs = 1.0/contraction_intervals
        delta_T = self.timeindex[1]-self.timeindex[0] # sampleperiod, assume equidistant sampling
        
        delta_ct = contraction_intervals
        err_freqs = delta_T/((delta_ct-delta_T)*delta_ct) 
        # estimate error: difference between maximal frequency betwen 2 peaks (fmax = 1/(delta_ct-1/delta_T)
        # (assuming an error in range of sampleperiod delta_T)
        # and detected frequency at sampling point (f= 1/delta_ct)
        
        return 60.0*freqs, 60.0*err_freqs #return in bpm
    
    def export_analysis(self, results_folder):
        #save peaks and peakanalysis to excel file:
        
        workbook_peaks = Workbook()
        sheet_peaks = workbook_peaks.active
        sheet_peaks.title = 'Analyzed peaks'
        sheet_peaks.append(['t [s]', 'motion [µm/s]', 'high/low'])         
        
        for peak in self.Peaks:
        
            time = self.timeindex[peak]
            motion = self.motion[peak]
            
            if peak in self.hipeaks:
                sheet_peaks.append([time, motion, "high"])
            elif peak in self.lopeaks:
                sheet_peaks.append([time, motion, "low"])
        
        sheet_analysis = workbook_peaks.create_sheet("Peak analysis")
        #ws.title = "New Title"
        ps = self.peakstatistics # introduced to shorten calls to variablename
        
        sheet_analysis.append(['value', 'mean', 'standard deviation', 'unit'])
        sheet_analysis.append(['maximum contraction', ps["max_contraction"], ps["max_contraction_std"], u'\xb5m/s'])
        sheet_analysis.append(['maximum relaxation', ps["max_relaxation"], ps["max_relaxation_std"], u'\xb5m/s'])
        sheet_analysis.append(['mean contraction interval', ps["contraction_interval_mean"], ps["contraction_interval_std"], 's'])
        sheet_analysis.append(['mean relaxation interval', ps["relaxation_interval_mean"], ps["relaxation_interval_std"], 's'])
        sheet_analysis.append(['mean contraction-relaxation interval', ps["contr_relax_interval_mean"], ps["contr_relax_interval_std"], 's'])
        sheet_analysis.append(['beating rate', ps["bpm_mean"], ps["bpm_std"], 'beats/min'])        

        sheet_analysis.column_dimensions['A'].width = 31
        sheet_analysis.column_dimensions['C'].width = 17
        
        sheet_kinetics = workbook_peaks.create_sheet("Kinetics",0)
        sheet_kinetics.append(['t [s]', 'motion [µm/s]'])

        for time, motion in zip(self.timeindex,self.motion):
            sheet_kinetics.append([time, motion])
        
        # quick workaround, interval calculation repeated, TODO: improve
        sheet_bpm_evol = workbook_peaks.create_sheet("bpm evolution")
        sheet_bpm_evol.append(['t [s]', 'f [bpm]', 'error [bpm]']) 
        # time indicates mean time of subsequent peaks which are used to calculate interval

        times_high = self.timeindex[self.hipeaks] if self.hipeaks != None else []
        contraction_intervals = np.diff(np.sort(times_high))
        freqs, errs = self.calc_bpm_evolution(contraction_intervals)
        meantimes = (times_high[:-1] + times_high[1:])/2
        for time, freq, err in zip(meantimes, freqs, errs):
            sheet_bpm_evol.append([time, freq, err])        

        results_folder.mkdir(parents = True, exist_ok = True) #TODO: check again, results_folder might not always be a path
        save_file = str(results_folder / 'Motionanalysis.xlsx')
        
        try:
            workbook_peaks.save(save_file)
        except PermissionError:
            print("can't access outputfile, "
                    "maybe it's still open?") 
                    # plot error (e.g. when called in cli), raise 
                    # needed such that gui-calls can catch error
            raise